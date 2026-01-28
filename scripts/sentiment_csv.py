import pandas as pd
import numpy as np
import re
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Téléchargement des ressources NLTK
import nltk
nltk.download('vader_lexicon', quiet=True)
nltk.download('stopwords', quiet=True)

from nltk.sentiment import SentimentIntensityAnalyzer
from nltk.corpus import stopwords
from textblob import TextBlob

# Initialiser l'analyseur de sentiment
sia = SentimentIntensityAnalyzer()

def clean_text(text):
    """Nettoyer le texte pour l'analyse"""
    if isinstance(text, str):
        # Convertir en minuscules
        text = text.lower()
        # Supprimer les URLs
        text = re.sub(r'http\S+|www\S+|https\S+', '', text, flags=re.MULTILINE)
        # Supprimer les mentions
        text = re.sub(r'@\w+', '', text)
        # Supprimer les hashtags
        text = re.sub(r'#\w+', '', text)
        # Supprimer les caractères spéciaux (garder les lettres, chiffres, espaces et accents français)
        text = re.sub(r'[^\w\sàâäéèêëîïôöùûüçÀÂÄÉÈÊËÎÏÔÖÙÛÜÇ]', ' ', text)
        # Supprimer les espaces multiples
        text = re.sub(r'\s+', ' ', text).strip()
        return text
    return ""

def analyze_sentiment_nltk(text):
    """Analyser le sentiment avec VADER"""
    cleaned_text = clean_text(text)
    scores = sia.polarity_scores(cleaned_text)
    
    # Déterminer la catégorie de sentiment
    if scores['compound'] >= 0.05:
        return 'positif', scores['compound']
    elif scores['compound'] <= -0.05:
        return 'négatif', scores['compound']
    else:
        return 'neutre', scores['compound']

def analyze_sentiment_textblob(text):
    """Analyser le sentiment avec TextBlob"""
    cleaned_text = clean_text(text)
    analysis = TextBlob(cleaned_text)
    
    # TextBlob retourne une polarité entre -1 et 1
    polarity = analysis.sentiment.polarity
    
    if polarity > 0.1:
        return 'positif', polarity
    elif polarity < -0.1:
        return 'négatif', polarity
    else:
        return 'neutre', polarity

def determine_final_sentiment(nltk_sentiment, textblob_sentiment, nltk_score, textblob_score):
    """Déterminer le sentiment final basé sur les deux méthodes"""
    
    # Si les deux méthodes sont d'accord
    if nltk_sentiment == textblob_sentiment:
        return nltk_sentiment, 'accord_total'
    
    # Si divergence, priorité à NLTK VADER (mieux pour réseaux sociaux)
    else:
        # Si NLTK est très certain (score élevé)
        if abs(nltk_score) > 0.5:
            return nltk_sentiment, 'priorité_nltk'
        # Si TextBlob est très certain
        elif abs(textblob_score) > 0.7:
            return textblob_sentiment, 'priorité_textblob'
        # Cas particulier: si un dit neutre et l'autre positif/négatif
        elif 'neutre' in [nltk_sentiment, textblob_sentiment]:
            non_neutre = nltk_sentiment if nltk_sentiment != 'neutre' else textblob_sentiment
            return non_neutre, 'neutre_contredit'
        # Sinon, neutre par défaut
        else:
            return 'neutre', 'divergence_resolue'

def analyze_and_classify_comments(df):
    """Analyser et classer les commentaires"""
    results = []
    
    print(f"Analyse et classification de {len(df)} commentaires...")
    
    for idx, row in df.iterrows():
        comment = row['Comment Text']
        
        # Analyse avec NLTK
        sentiment_nltk, score_nltk = analyze_sentiment_nltk(comment)
        
        # Analyse avec TextBlob
        sentiment_tb, score_tb = analyze_sentiment_textblob(comment)
        
        # Décision finale
        final_sentiment, decision_reason = determine_final_sentiment(
            sentiment_nltk, sentiment_tb, score_nltk, score_tb
        )
        
        # Format de date pour Excel
        date_str = ""
        if 'Post Date-Time' in row and pd.notna(row['Post Date-Time']):
            try:
                date_str = pd.to_datetime(row['Post Date-Time']).strftime('%Y-%m-%d %H:%M:%S')
            except:
                date_str = str(row['Post Date-Time'])
        
        results.append({
            'ID': idx + 1,
            'Commentaire': comment,
            'Auteur': row['Author'] if 'Author' in row else 'Inconnu',
            'Date': date_str,
            'Sentiment_NLTK': sentiment_nltk,
            'Score_NLTK': round(score_nltk, 3),
            'Sentiment_TextBlob': sentiment_tb,
            'Score_TextBlob': round(score_tb, 3),
            'Sentiment_Final': final_sentiment,
            'Raison_Decision': decision_reason,
            'Confiance': min(abs(score_nltk), abs(score_tb)) * 100
        })
        
        # Afficher la progression
        if (idx + 1) % 50 == 0:
            print(f"  {idx + 1}/{len(df)} commentaires traités...")
    
    return pd.DataFrame(results)

def create_excel_with_sheets(df_sentiments, output_filename='commentaires_classes.xlsx'):
    """Créer un fichier Excel avec des onglets séparés pour chaque sentiment"""
    
    # Séparer les commentaires par sentiment
    df_positif = df_sentiments[df_sentiments['Sentiment_Final'] == 'positif'].copy()
    df_neutre = df_sentiments[df_sentiments['Sentiment_Final'] == 'neutre'].copy()
    df_negatif = df_sentiments[df_sentiments['Sentiment_Final'] == 'négatif'].copy()
    
    # Trier par score de confiance (décroissant)
    df_positif = df_positif.sort_values('Score_NLTK', ascending=False)
    df_negatif = df_negatif.sort_values('Score_NLTK', ascending=True)  # Les plus négatifs d'abord
    df_neutre = df_neutre.sort_values('Confiance', ascending=True)  # Les moins confiants d'abord
    
    # Réinitialiser les index
    df_positif.reset_index(drop=True, inplace=True)
    df_neutre.reset_index(drop=True, inplace=True)
    df_negatif.reset_index(drop=True, inplace=True)
    
    # Mettre à jour les IDs
    df_positif['ID'] = range(1, len(df_positif) + 1)
    df_neutre['ID'] = range(1, len(df_neutre) + 1)
    df_negatif['ID'] = range(1, len(df_negatif) + 1)
    
    # Créer un writer Excel
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # Écrire chaque dataframe dans un onglet séparé
        df_positif.to_excel(writer, sheet_name='Commentaires_Positifs', index=False)
        df_neutre.to_excel(writer, sheet_name='Commentaires_Neutres', index=False)
        df_negatif.to_excel(writer, sheet_name='Commentaires_Négatifs', index=False)
        
        # Créer un onglet de résumé
        summary_data = {
            'Catégorie': ['Positifs', 'Neutres', 'Négatifs', 'Total'],
            'Nombre': [len(df_positif), len(df_neutre), len(df_negatif), len(df_sentiments)],
            'Pourcentage': [
                f"{(len(df_positif)/len(df_sentiments)*100):.1f}%",
                f"{(len(df_neutre)/len(df_sentiments)*100):.1f}%",
                f"{(len(df_negatif)/len(df_sentiments)*100):.1f}%",
                "100%"
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Résumé', index=False)
        
        # Onglet complet avec tous les commentaires
        df_sentiments.to_excel(writer, sheet_name='Tous_Commentaires', index=False)
    
    return df_positif, df_neutre, df_negatif

def format_excel_columns(output_filename='commentaires_classes.xlsx'):
    """Formater les colonnes Excel pour une meilleure lisibilité"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Charger le workbook
        wb = load_workbook(output_filename)
        
        # Définir les styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Formater chaque feuille
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ajuster la largeur des colonnes
            if sheet_name == 'Commentaires_Positifs':
                column_widths = {'A': 5, 'B': 60, 'C': 20, 'D': 18, 'E': 12, 'F': 10, 
                                'G': 12, 'H': 10, 'I': 12, 'J': 15, 'K': 10}
                # Fond vert pour les positifs
                data_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif sheet_name == 'Commentaires_Négatifs':
                column_widths = {'A': 5, 'B': 60, 'C': 20, 'D': 18, 'E': 12, 'F': 10, 
                                'G': 12, 'H': 10, 'I': 12, 'J': 15, 'K': 10}
                # Fond rouge clair pour les négatifs
                data_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif sheet_name == 'Commentaires_Neutres':
                column_widths = {'A': 5, 'B': 60, 'C': 20, 'D': 18, 'E': 12, 'F': 10, 
                                'G': 12, 'H': 10, 'I': 12, 'J': 15, 'K': 10}
                # Fond jaune clair pour les neutres
                data_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                column_widths = {'A': 5, 'B': 60, 'C': 20, 'D': 18, 'E': 12, 'F': 10, 
                                'G': 12, 'H': 10, 'I': 12, 'J': 15, 'K': 10}
                data_fill = None
            
            # Appliquer les largeurs de colonnes
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Formater l'en-tête
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border
            
            # Formater les données
            max_row = ws.max_row
            max_col = ws.max_column
            
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = thin_border
                    
                    # Appliquer le fond coloré si disponible
                    if data_fill and sheet_name in ['Commentaires_Positifs', 'Commentaires_Négatifs', 'Commentaires_Neutres']:
                        cell.fill = data_fill
            
            # Geler la première ligne (en-tête)
            ws.freeze_panes = 'A2'
        
        # Sauvegarder les modifications
        wb.save(output_filename)
        print(f"✅ Formatage Excel terminé : {output_filename}")
        
    except ImportError:
        print("⚠️  openpyxl non installé, formatage Excel simplifié")
        print("   Installez avec: pip install openpyxl")
    except Exception as e:
        print(f"⚠️  Erreur lors du formatage Excel: {str(e)}")

def generate_statistics_report(df_sentiments, df_positif, df_neutre, df_negatif):
    """Générer un rapport statistique"""
    
    print("\n" + "="*80)
    print("📊 RAPPORT STATISTIQUE DE CLASSIFICATION")
    print("="*80)
    
    total = len(df_sentiments)
    
    print(f"\n📈 RÉPARTITION DES SENTIMENTS:")
    print("-"*40)
    print(f"  • Commentaires positifs : {len(df_positif):>4} ({len(df_positif)/total*100:.1f}%)")
    print(f"  • Commentaires neutres  : {len(df_neutre):>4} ({len(df_neutre)/total*100:.1f}%)")
    print(f"  • Commentaires négatifs : {len(df_negatif):>4} ({len(df_negatif)/total*100:.1f}%)")
    print(f"  • TOTAL                : {total:>4} (100%)")
    
    print(f"\n🎯 QUALITÉ DE L'ANALYSE:")
    print("-"*40)
    
    # Calcul de la concordance
    concordance = (df_sentiments['Sentiment_NLTK'] == df_sentiments['Sentiment_TextBlob']).mean() * 100
    accord_total = (df_sentiments['Raison_Decision'] == 'accord_total').sum()
    
    print(f"  • Concordance NLTK-TextBlob : {concordance:.1f}%")
    print(f"  • Décisions unanimes        : {accord_total} ({accord_total/total*100:.1f}%)")
    
    # Analyse par type de décision
    decision_counts = df_sentiments['Raison_Decision'].value_counts()
    print(f"\n🔍 RÉPARTITION DES DÉCISIONS:")
    print("-"*40)
    for decision, count in decision_counts.items():
        print(f"  • {decision:<20} : {count:>4} ({count/total*100:.1f}%)")
    
    # Top 5 des commentaires les plus positifs
    if len(df_positif) > 0:
        print(f"\n🏆 TOP 5 COMMENTAIRES LES PLUS POSITIFS:")
        print("-"*40)
        for i, row in df_positif.head(5).iterrows():
            print(f"  {i+1}. Score: {row['Score_NLTK']:.3f} - {row['Commentaire'][:80]}...")
    
    # Top 5 des commentaires les plus négatifs
    if len(df_negatif) > 0:
        print(f"\n⚠️  TOP 5 COMMENTAIRES LES PLUS NÉGATIFS:")
        print("-"*40)
        for i, row in df_negatif.head(5).iterrows():
            print(f"  {i+1}. Score: {row['Score_NLTK']:.3f} - {row['Commentaire'][:80]}...")
    
    print(f"\n💾 FICHIERS GÉNÉRÉS:")
    print("-"*40)
    print("  1. commentaires_classes.xlsx - Fichier Excel avec onglets séparés")
    print("     • Commentaires_Positifs - Tous les commentaires positifs")
    print("     • Commentaires_Neutres  - Tous les commentaires neutres")
    print("     • Commentaires_Négatifs - Tous les commentaires négatifs")
    print("     • Résumé - Statistiques globales")
    print("     • Tous_Commentaires - Tous les commentaires avec analyse")
    print("  2. commentaires_detailles.csv - Fichier CSV complet")
    print("  3. statistiques.txt - Rapport détaillé")

def main():
    """Fonction principale"""
    
    # Chemin vers votre fichier CSV
    csv_path = r"C:\Users\Sy Savane Idriss\project_sentiment_fb\data\fb_comment.csv"
    
    print("="*80)
    print("🚀 DÉMARRAGE DE LA CLASSIFICATION DES COMMENTAIRES")
    print("="*80)
    print(f"📂 Lecture du fichier: {csv_path}")
    
    try:
        # Lecture du fichier CSV
        df = pd.read_csv(csv_path, encoding='utf-8')
        
        print(f"✅ Fichier chargé avec succès")
        print(f"   • Nombre de commentaires: {len(df):,}")
        print(f"   • Colonnes disponibles: {list(df.columns)}")
        
        # Vérifier les colonnes nécessaires
        if 'Comment Text' not in df.columns:
            raise ValueError("La colonne 'Comment Text' est introuvable dans le fichier CSV")
        
        # Analyse et classification
        print("\n🔍 Analyse et classification en cours...")
        df_sentiments = analyze_and_classify_comments(df)
        
        # Création du fichier Excel
        print("\n📊 Création du fichier Excel...")
        df_positif, df_neutre, df_negatif = create_excel_with_sheets(df_sentiments)
        
        # Formater le fichier Excel
        format_excel_columns()
        
        # Sauvegarde CSV complémentaire
        csv_output = 'commentaires_detailles.csv'
        df_sentiments.to_csv(csv_output, index=False, encoding='utf-8-sig')
        print(f"✅ Fichier CSV détaillé sauvegardé : {csv_output}")
        
        # Générer le rapport statistique
        generate_statistics_report(df_sentiments, df_positif, df_neutre, df_negatif)
        
        # Sauvegarder le rapport dans un fichier texte
        with open('statistiques.txt', 'w', encoding='utf-8') as f:
            import sys
            old_stdout = sys.stdout
            sys.stdout = f
            generate_statistics_report(df_sentiments, df_positif, df_neutre, df_negatif)
            sys.stdout = old_stdout
        
        print("\n" + "="*80)
        print("✅ CLASSIFICATION TERMINÉE AVEC SUCCÈS!")
        print("="*80)
        
        print(f"\n🎯 RÉSULTATS FINAUX:")
        print(f"   • {len(df_positif)} commentaires positifs classés")
        print(f"   • {len(df_neutre)} commentaires neutres classés")
        print(f"   • {len(df_negatif)} commentaires négatifs classés")
        
        print(f"\n📁 OUVREZ LE FICHIER : commentaires_classes.xlsx")
        print("   pour voir les commentaires classés par catégorie!")
        
    except FileNotFoundError:
        print(f"❌ ERREUR: Fichier non trouvé à l'emplacement: {csv_path}")
        print("   Vérifiez le chemin du fichier.")
    except Exception as e:
        print(f"❌ ERREUR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Installation des dépendances si nécessaire
    try:
        from openpyxl import Workbook
    except ImportError:
        print("📦 Installation des dépendances supplémentaires...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        print("✅ Dépendances installées avec succès")
    
    main()